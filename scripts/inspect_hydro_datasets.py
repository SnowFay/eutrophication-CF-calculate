#!/usr/bin/env python3
"""Generate a schema/attribute inventory for the HydroSHEDS datasets in this repo."""
from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import fiona
import xarray as xr

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - openpyxl is an optional helper
    load_workbook = None  # type: ignore

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_ROOT = BASE_DIR / "data"
HYDRO_DATA_DIR = DATA_ROOT / "HydroSHEDS data"
WATER_DATA_DIR = DATA_ROOT / "water use data"
DEFAULT_OUTPUT = BASE_DIR / "materials" / "dataset_field_inventory.md"
VARIABLE_METADATA_PATH = BASE_DIR / "materials" / "hydroatlas_variable_metadata.json"


@dataclass(frozen=True)
class DatasetConfig:
    """Describe where a dataset lives and how to iterate through its layers."""

    key: str
    name: str
    path: Path
    layers: Sequence[Optional[str]]
    legend_workbook: Optional[Path] = None
    documentation: Sequence[Path] = ()
    data_format: str = "vector"
    extra_field_hints: Dict[str, str] = field(default_factory=dict)

    def available_layers(self) -> Sequence[Optional[str]]:
        return self.layers


DATASETS: Sequence[DatasetConfig] = (
    DatasetConfig(
        key="hydrorivers_asia",
        name="HydroRIVERS v1.0 – Asia subset",
        path=HYDRO_DATA_DIR
        / "HydroRIVERS_v10_as_shp/HydroRIVERS_v10_as_shp/HydroRIVERS_v10_as.shp",
        layers=[None],
        documentation=(
            HYDRO_DATA_DIR / "HydroRIVERS_v10_as_shp/HydroRIVERS_TechDoc_v10.pdf",
        ),
    ),
    DatasetConfig(
        key="hydrolakes_polys",
        name="HydroLAKES v1.0 – polygons",
        path=HYDRO_DATA_DIR
        / "HydroLAKES_polys_v10_shp/HydroLAKES_polys_v10_shp/HydroLAKES_polys_v10.shp",
        layers=[None],
        documentation=(
            HYDRO_DATA_DIR / "HydroLAKES_polys_v10_shp/HydroLAKES_TechDoc_v10.pdf",
        ),
    ),
    DatasetConfig(
        key="hydrolakes_points",
        name="HydroLAKES v1.0 – centroids",
        path=HYDRO_DATA_DIR
        / "HydroLAKES_points_v10_shp/HydroLAKES_points_v10_shp/HydroLAKES_points_v10.shp",
        layers=[None],
        documentation=(
            HYDRO_DATA_DIR / "HydroLAKES_points_v10_shp/HydroLAKES_TechDoc_v10.pdf",
        ),
    ),
    DatasetConfig(
        key="hydrobasins_lake_level10",
        name="HydroBASINS level 10 lake basins",
        path=HYDRO_DATA_DIR / "hybas_lake_as_lev10_v1c/hybas_lake_as_lev10_v1c.shp",
        layers=[None],
        documentation=(
            HYDRO_DATA_DIR / "hybas_lake_as_lev10_v1c/HydroBASINS_TechDoc_v1c.pdf",
        ),
    ),
    DatasetConfig(
        key="hydrobasins_pourpoints",
        name="HydroBASINS level 10 pour points",
        path=HYDRO_DATA_DIR / "hybas_pour_lev10_v1_shp/hybas_pour_lev10_v1.shp",
        layers=[None],
        documentation=(
            HYDRO_DATA_DIR / "hybas_pour_lev10_v1_shp/HydroBASINS_PourPoint_TechDoc_v1.pdf",
        ),
    ),
    DatasetConfig(
        key="riveratlas",
        name="RiverATLAS v1.0 (global reaches)",
        path=HYDRO_DATA_DIR / "RiverATLAS_Data_v10.gdb/RiverATLAS_v10.gdb",
        layers=["RiverATLAS_v10"],
        legend_workbook=
        HYDRO_DATA_DIR / "RiverATLAS_Data_v10.gdb/HydroATLAS_v10_Legends.xlsx",
        documentation=(
            HYDRO_DATA_DIR / "RiverATLAS_Data_v10.gdb/RiverATLAS_Catalog_v10.pdf",
            HYDRO_DATA_DIR / "RiverATLAS_Data_v10.gdb/HydroATLAS_TechDoc_v10.pdf",
        ),
    ),
    DatasetConfig(
        key="basinatlas",
        name="BasinATLAS v1.0 (HydroBASINS derived attributes)",
        path=HYDRO_DATA_DIR / "BasinATLAS_Data_v10.gdb/BasinATLAS_v10.gdb",
        layers=[f"BasinATLAS_v10_lev{level:02d}" for level in range(1, 13)],
        legend_workbook=
        HYDRO_DATA_DIR / "BasinATLAS_Data_v10.gdb/HydroATLAS_v10_Legends.xlsx",
        documentation=(
            HYDRO_DATA_DIR / "BasinATLAS_Data_v10.gdb/BasinATLAS_Catalog_v10.pdf",
            HYDRO_DATA_DIR / "BasinATLAS_Data_v10.gdb/HydroATLAS_TechDoc_v10.pdf",
        ),
    ),
    DatasetConfig(
        key="lakeatlas",
        name="LakeATLAS v1.0",
        path=HYDRO_DATA_DIR / "LakeATLAS_Data_v10.gdb/LakeATLAS_v10.gdb",
        layers=["LakeATLAS_v10_pol", "LakeATLAS_v10_pnt"],
        legend_workbook=
        HYDRO_DATA_DIR / "LakeATLAS_Data_v10.gdb/HydroATLAS_v10_Legends.xlsx",
        documentation=(
            HYDRO_DATA_DIR / "LakeATLAS_Data_v10.gdb/LakeATLAS_Catalog_v10.pdf",
            HYDRO_DATA_DIR / "LakeATLAS_Data_v10.gdb/HydroATLAS_TechDoc_v10_1.pdf",
        ),
    ),
    DatasetConfig(
        key="hswud_dom",
        name="HSWUD domestic water use (monthly, 1965–2022)",
        path=WATER_DATA_DIR / "27610524/HSWUD_dom.nc",
        layers=[None],
        documentation=(WATER_DATA_DIR / "27610524/state.md",),
        data_format="netcdf",
        extra_field_hints={
            "dom": "Monthly domestic water use withdrawals (10^8 m³) from HSWUD dataset.",
        },
    ),
    DatasetConfig(
        key="hswud_ele",
        name="HSWUD electricity sector water use",
        path=WATER_DATA_DIR / "27610524/HSWUD_ele.nc",
        layers=[None],
        documentation=(WATER_DATA_DIR / "27610524/state.md",),
        data_format="netcdf",
        extra_field_hints={
            "ele": "Monthly thermoelectric water withdrawals (10^8 m³) from HSWUD dataset.",
        },
    ),
    DatasetConfig(
        key="hswud_irr",
        name="HSWUD irrigation water use",
        path=WATER_DATA_DIR / "27610524/HSWUD_irr_.nc",
        layers=[None],
        documentation=(WATER_DATA_DIR / "27610524/state.md",),
        data_format="netcdf",
        extra_field_hints={
            "irr": "Monthly irrigation water withdrawals (10^8 m³) from HSWUD dataset.",
        },
    ),
    DatasetConfig(
        key="hswud_manu",
        name="HSWUD manufacturing water use",
        path=WATER_DATA_DIR / "27610524/HSWUD_manu.nc",
        layers=[None],
        documentation=(WATER_DATA_DIR / "27610524/state.md",),
        data_format="netcdf",
        extra_field_hints={
            "manu": "Monthly manufacturing water use withdrawals (10^8 m³) from HSWUD dataset.",
        },
    ),
)


# -- Field description helpers -------------------------------------------------

EXACT_FIELD_HINTS: Dict[str, str] = {
    "HYRIV_ID": "Permanent HydroRIVERS/HydroATLAS river reach identifier.",
    "NEXT_DOWN": "HYRIV_ID of the immediate downstream reach (0 = ocean outlet).",
    "MAIN_RIV": "Identifier for the main river stem that the reach belongs to.",
    "LENGTH_KM": "Reach length in kilometers (planform).",
    "DIST_DN_KM": "Distance from the reach outlet to the ocean (km).",
    "DIST_UP_KM": "Distance from headwaters to the reach inlet (km).",
    "CATCH_SKM": "Surface area (km²) of the local catchment draining directly to the reach.",
    "UPLAND_SKM": "Surface area (km²) of the entire upstream drainage area.",
    "ENDORHEIC": "1 if the reach drains to an inland sink, 0 otherwise.",
    "DIS_AV_CMS": "Long-term average discharge in cubic meters per second at the reach outlet.",
    "ORD_STRA": "Strahler stream order of the reach.",
    "ORD_CLAS": "HydroSHEDS level-based stream order.",
    "ORD_FLOW": "Shreve (flow-based) stream order.",
    "HYBAS_L12": "Identifier of the intersecting HydroBASINS level-12 polygon.",
    "HYBAS_ID": "Permanent HydroBASINS polygon identifier at the dataset's level.",
    "NEXT_SINK": "HYBAS_ID of the downstream sink (endorheic target).",
    "MAIN_BAS": "Identifier of the encompassing main basin (Pfafstetter coding).",
    "DIST_SINK": "Flow distance (km) from the basin outlet to its terminal sink.",
    "DIST_MAIN": "Flow distance (km) from the basin outlet to the main river mouth.",
    "SUB_AREA": "Area (km²) of the polygon between nested HydroBASINS levels.",
    "UP_AREA": "Total upstream area (km²) draining through the polygon outlet.",
    "PFAF_ID": "Pfafstetter hierarchical basin code.",
    "SIDE": "Pfafstetter side code (left/right).",
    "LAKE": "HydroLAKES (Hylak_id) identifier for lake polygons matched to the basin.",
    "COAST": "1 if the basin drains directly to the ocean.",
    "ORDER": "Topological order/index used for sorting HydroBASINS polygons.",
    "SORT": "Pre-calculated sort key for topological traversals.",
    "Hylak_id": "Primary HydroLAKES identifier.",
    "Lake_name": "Lake/reservoir name (if assigned).",
    "Country": "Primary ISO-3 country code (may be multi-valued).",
    "Continent": "Continent tag used by HydroLAKES.",
    "Poly_src": "Source dataset for the polygon geometry.",
    "Lake_type": "HydroLAKES classification (natural lake, reservoir, etc.).",
    "Grand_id": "GRanD database identifier (if the lake is a registered reservoir).",
    "Lake_area": "Open water surface area (km²).",
    "Shore_len": "Shoreline length (km).",
    "Shore_dev": "Shoreline development index (dimensionless).",
    "Vol_total": "Total lake/reservoir volume (km³).",
    "Vol_res": "Managed reservoir volume component (km³).",
    "Vol_src": "Source authority for the volume estimate.",
    "Depth_avg": "Mean depth (m).",
    "Dis_avg": "Estimated mean annual outflow/discharge (m³/s).",
    "Res_time": "Mean residence time (years).",
    "Elevation": "Lake surface elevation (m a.s.l.).",
    "Slope_100": "Mean 100 m buffer slope (degrees).",
    "Wshd_area": "Total watershed area feeding the lake (km²).",
    "Pour_long": "Longitude of the modeled pour point (degrees).",
    "Pour_lat": "Latitude of the modeled pour point (degrees).",
    "Shape_Length": "Geometry perimeter/length reported by the data source CRS.",
    "Shape_Area": "Geometry area reported by the data source CRS.",
}

VARIABLE_HINTS: Dict[str, str] = {
    "tmp": "Near-surface air temperature",
    "pre": "Precipitation accumulation",
    "pet": "Potential evapotranspiration",
    "aet": "Actual evapotranspiration",
    "run": "Runoff depth",
    "dis": "Discharge volume",
    "inu": "Inundation or flooding fraction",
    "lka": "Lake fraction (GLWD)",
    "lkv": "Lake/reservoir volume estimate",
    "rev": "Reservoir volume",
    "dor": "Degree of regulation",
    "ria": "Riverine inundated area",
    "riv": "River network channel density",
    "gwt": "Groundwater table depth",
    "ele": "Elevation",
    "slp": "Surface slope",
    "sgr": "Stream gradient",
    "clz": "Köppen-Geiger climate zone code",
    "cls": "Global Environmental Stratification class",
    "snw": "Snow cover fraction",
    "glc": "GLC2000 land-cover shares",
    "pnv": "Potential natural vegetation shares",
    "wet": "GLWD wetland shares",
    "for": "Tree/forest cover share",
    "crp": "Cropland share",
    "pst": "Pasture share",
    "ire": "Irrigated area share",
    "gla": "Glacier cover share",
    "prm": "Permanent water share",
    "pac": "Rainfed/paddy agriculture share",
    "tbi": "Terrestrial biome class",
    "tec": "Terrestrial ecoregion class",
    "fmh": "Freshwater major habitat type",
    "fec": "Freshwater ecoregion class",
    "cly": "Topsoil clay fraction",
    "slt": "Topsoil silt fraction",
    "snd": "Topsoil sand fraction",
    "soc": "Soil organic carbon",
    "swc": "Soil water content",
    "lit": "Lithology class",
    "kar": "Karst proportion",
    "ero": "Soil erosion rate",
    "pop": "Population count",
    "ppd": "Population density",
    "urb": "Urban land fraction",
    "nli": "Nighttime lights index",
    "rdd": "Road density",
    "hft": "Human footprint index",
    "gad": "Administrative identifier (GAUL / GADM cross-walk)",
    "gdp": "GDP (PPP) per capita",
    "hdi": "UNDP Human Development Index",
}

UNIT_HINTS: Dict[str, str] = {
    "dc": "°C",
    "mm": "mm",
    "cms": "m³/s",
    "m3": "m³",
    "pc": "percent of the referenced area",
    "mc": "million cubic meters",
    "ha": "hectares",
    "ix": "dimensionless index",
    "cm": "centimeters",
    "mt": "meters",
    "dg": "degrees",
    "dk": "m/km (slope or gradient)",
    "th": "tonnes per hectare",
    "pk": "persons per km²",
    "ud": "US dollars (constant PPP, per capita)",
    "tc": "total channel length (km)",
}

AREA_CODES: Dict[str, str] = {
    "c": "local catchment aggregate",
    "u": "entire upstream area aggregate",
    "s": "HydroBASINS polygon aggregate",
    "p": "pour-point statistic (network outlet)",
    "r": "reach-length weighted statistic",
}

MONTH_NAMES = {
    f"{i:02d}": name
    for i, name in enumerate(
        (
            "January",
            "February",
            "March",
            "April",
            "May",
            "June",
            "July",
            "August",
            "September",
            "October",
            "November",
            "December",
        ),
        start=1,
    )
}

METRIC_HINTS: Dict[str, str] = {
    "yr": "long-term annual mean",
    "mn": "minimum of long-term monthly climatology",
    "mx": "maximum of long-term monthly climatology",
    "av": "area-weighted average",
    "su": "total sum/extent",
    "se": "fractional share/extent",
    "lt": "latest available snapshot (~2010 baseline)",
    "mj": "dominant (majority) class ID",
    "g1": "GLWD class group 1 (permanent wetlands)",
    "g2": "GLWD class group 2 (seasonal wetlands)",
    "93": "value circa 1993",
    "09": "value circa 2009",
    "pyr": "annual discharge total at the pour point",
    "pmn": "minimum monthly discharge",
    "pmx": "maximum monthly discharge",
    "pva": "coefficient of variation of monthly discharge",
    "clt": "latest climatology (2000s)",
    "ult": "latest climatology (upstream aggregate)",
    "umn": "upstream minimum",
    "umx": "upstream maximum",
    "uav": "upstream average",
    "uyr": "upstream annual mean",
    "usu": "upstream total",
    "use": "upstream share",
    "cyr": "local annual mean",
    "csu": "local total",
    "cse": "local share",
    "cmn": "local minimum",
    "cmx": "local maximum",
    "cav": "local average",
    "sav": "sub-basin average",
    "syr": "sub-basin annual mean",
    "ssu": "sub-basin sum",
    "sse": "sub-basin share",
    "smn": "sub-basin minimum",
    "smx": "sub-basin maximum",
}

CLASS_SHEETS = {
    "clz_cl",
    "cls_cl",
    "glc_cl",
    "pnv_cl",
    "wet_cl",
    "tbi_cl",
    "tec_cl",
    "fmh_cl",
    "fec_cl",
    "lit_cl",
    "gad_id",
}

FALLBACK_DESCRIPTION = "See accompanying technical documentation for semantic details."


def _load_variable_hint_overrides() -> Dict[str, str]:
    if not VARIABLE_METADATA_PATH.exists():
        return {}
    try:
        with VARIABLE_METADATA_PATH.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
    except Exception:
        return {}
    return {
        key.lower(): str(value)
        for key, value in data.items()
        if isinstance(key, str)
    }


VARIABLE_HINTS.update(_load_variable_hint_overrides())


@lru_cache(maxsize=None)
def legend_sheets(workbook: Path) -> Sequence[str]:
    if not load_workbook or not workbook.exists():
        return []
    wb = load_workbook(workbook, read_only=True, data_only=True)
    try:
        return tuple(sheet.lower() for sheet in wb.sheetnames)
    finally:
        wb.close()


class FieldDescriber:
    """Generate human-readable hints for attribute names."""

    def __init__(
        self,
        legend_workbook: Optional[Path] = None,
        extra_hints: Optional[Dict[str, str]] = None,
    ):
        self.legend_workbook = legend_workbook
        self._legend_cache = None
        self.extra_hints = {k.lower(): v for k, v in (extra_hints or {}).items()}

    def describe(self, name: str) -> str:
        base = name.strip()
        if base in EXACT_FIELD_HINTS:
            return EXACT_FIELD_HINTS[base]
        lowered = base.lower()
        if lowered in self.extra_hints:
            return self.extra_hints[lowered]
        class_hint = self._classification_hint(lowered)
        if class_hint:
            return class_hint
        parts = lowered.split("_")
        if len(parts) >= 3:
            var, unit, token = parts[:3]
            desc = self._compose_description(var, unit, token)
            if desc:
                return desc
        if len(parts) == 2:
            desc = self._compose_description(parts[0], parts[1], None)
            if desc:
                return desc
        return FALLBACK_DESCRIPTION

    def _compose_description(
        self, var: str, unit: Optional[str], token: Optional[str]
    ) -> Optional[str]:
        segments: List[str] = []
        if var in VARIABLE_HINTS:
            segments.append(VARIABLE_HINTS[var])
        if unit and unit in UNIT_HINTS:
            segments.append(f"units: {UNIT_HINTS[unit]}")
        if token:
            scope = self._decode_token(token)
            if scope:
                segments.append(scope)
        if segments:
            return "; ".join(segments)
        return None

    def _decode_token(self, token: str) -> Optional[str]:
        token = token.lower()
        if not token:
            return None
        area_desc = None
        metric_code = token
        if token[0] in AREA_CODES:
            area_desc = AREA_CODES[token[0]]
            metric_code = token[1:]
        metric_desc = self._metric_description(metric_code)
        if area_desc and metric_desc:
            return f"{area_desc}; {metric_desc}"
        if area_desc:
            return area_desc
        return metric_desc

    def _metric_description(self, metric_code: str) -> Optional[str]:
        if not metric_code:
            return None
        if metric_code in METRIC_HINTS:
            return METRIC_HINTS[metric_code]
        if metric_code in MONTH_NAMES:
            return f"climatological mean for {MONTH_NAMES[metric_code]}"
        if metric_code.startswith("g") and metric_code[1:].isdigit():
            return "class group {} (see legend workbook)".format(metric_code[1:])
        if metric_code.isdigit():
            return f"classification bin {metric_code} (see legend workbook)"
        return None

    def _classification_hint(self, lowered_name: str) -> Optional[str]:
        if "_cl" not in lowered_name:
            return None
        prefix = lowered_name.split("_cl")[0]
        sheet_name = f"{prefix}_cl"
        if sheet_name not in CLASS_SHEETS:
            return None
        if self.legend_workbook and self._sheet_exists(sheet_name):
            return (
                f"Categorical codes; decode via sheet '{sheet_name}' in "
                f"{self.legend_workbook.name}."
            )
        return "Categorical codes; see HydroATLAS legend workbook."

    def _sheet_exists(self, sheet: str) -> bool:
        if not self.legend_workbook:
            return False
        if self._legend_cache is None:
            self._legend_cache = set(legend_sheets(self.legend_workbook))
        return sheet.lower() in self._legend_cache


@dataclass
class FieldSummary:
    name: str
    type_label: str
    description: str


@dataclass
class LayerSummary:
    dataset: DatasetConfig
    layer_name: Optional[str]
    path: Path
    driver: str
    geometry: str
    crs: Optional[str]
    feature_count: Optional[int]
    fields: Optional[List[FieldSummary]]
    schema_reference: Optional[str]
    extra_lines: Sequence[str] = ()

    @property
    def display_name(self) -> str:
        if self.layer_name:
            return self.layer_name
        return self.path.stem


class DatasetAnalyzer:
    """Collect schema information for every layer of a dataset."""

    def __init__(self, config: DatasetConfig):
        self.config = config
        self.describer = FieldDescriber(
            config.legend_workbook, extra_hints=config.extra_field_hints
        )
        self._schema_refs: Dict[Tuple[Tuple[str, str], ...], str] = {}

    def summarize(self) -> List[LayerSummary]:
        summaries: List[LayerSummary] = []
        for layer_name in self.config.available_layers():
            summaries.append(self._summarize_layer(layer_name))
        return summaries

    def _summarize_layer(self, layer_name: Optional[str]) -> LayerSummary:
        if self.config.data_format == "vector":
            return self._summarize_vector_layer(layer_name)
        if self.config.data_format == "netcdf":
            return self._summarize_netcdf_layer(layer_name)
        raise ValueError(f"Unsupported data_format: {self.config.data_format}")

    def _summarize_vector_layer(self, layer_name: Optional[str]) -> LayerSummary:
        path = self.config.path
        if not path.exists():
            raise FileNotFoundError(f"Missing dataset: {path}")
        with fiona.Env():
            with fiona.open(path, layer=layer_name) as src:
                schema_items = tuple(src.schema["properties"].items())
                schema_ref = self._schema_refs.get(schema_items)
                fields: Optional[List[FieldSummary]] = None
                if schema_ref is None:
                    fields = [
                        FieldSummary(
                            name=name,
                            type_label=str(dtype),
                            description=self.describer.describe(name),
                        )
                        for name, dtype in schema_items
                    ]
                    schema_ref = self.config.key + ":" + (layer_name or path.stem)
                    self._schema_refs[schema_items] = schema_ref
                return LayerSummary(
                    dataset=self.config,
                    layer_name=layer_name,
                    path=path,
                    driver=src.driver,
                    geometry=src.schema.get("geometry"),
                    crs=self._format_crs(src),
                    feature_count=self._safe_len(src),
                    fields=fields,
                    schema_reference=None if fields else schema_ref,
                    extra_lines=(),
                )

    def _summarize_netcdf_layer(self, layer_name: Optional[str]) -> LayerSummary:
        path = self.config.path
        if not path.exists():
            raise FileNotFoundError(f"Missing dataset: {path}")
        extra_lines: List[str] = []
        fields: List[FieldSummary] = []
        with xr.open_dataset(path) as ds:
            if ds.sizes:
                dims_desc = ", ".join(f"{dim}={size}" for dim, size in ds.sizes.items())
                extra_lines.append(f"Dimensions: {dims_desc}")
            if ds.attrs:
                for key, value in ds.attrs.items():
                    extra_lines.append(f"Global attr {key}: {value}")
            for var_name, data_array in ds.data_vars.items():
                dims_text = ", ".join(data_array.dims)
                type_label = f"{data_array.dtype} dims=({dims_text})"
                description = self._describe_data_array(var_name, data_array)
                fields.append(FieldSummary(var_name, type_label, description))
        return LayerSummary(
            dataset=self.config,
            layer_name=layer_name,
            path=path,
            driver="NetCDF",
            geometry=None,
            crs=None,
            feature_count=None,
            fields=fields,
            schema_reference=None,
            extra_lines=tuple(extra_lines),
        )

    def _describe_data_array(self, name: str, data_array) -> str:
        desc_parts: List[str] = []
        base_desc = self.describer.describe(name)
        if base_desc and base_desc != FALLBACK_DESCRIPTION:
            desc_parts.append(base_desc)
        attr_desc = data_array.attrs.get("description") or data_array.attrs.get("long_name")
        if attr_desc:
            desc_parts.append(str(attr_desc))
        units = data_array.attrs.get("units")
        if units:
            desc_parts.append(f"units: {units}")
        if not desc_parts:
            return FALLBACK_DESCRIPTION
        return "; ".join(desc_parts)

    @staticmethod
    def _safe_len(collection) -> Optional[int]:
        try:
            return len(collection)
        except Exception:  # pragma: no cover
            return None

    @staticmethod
    def _format_crs(src) -> Optional[str]:
        crs = src.crs_wkt or src.crs
        if not crs:
            return None
        if isinstance(crs, dict) and "init" in crs:
            return crs["init"]
        return str(crs)


# -- Reporting ----------------------------------------------------------------

def write_markdown_report(layer_summaries: Sequence[LayerSummary], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    by_dataset: Dict[str, List[LayerSummary]] = {}
    for summary in layer_summaries:
        by_dataset.setdefault(summary.dataset.key, []).append(summary)

    lines: List[str] = [
        "# HydroSHEDS / HydroATLAS dataset inventory",
        "",
        f"Generated {timestamp}",
        "",
        "This file is auto-created by `scripts/inspect_hydro_datasets.py`.",
        "",
    ]

    for cfg in DATASETS:
        group = by_dataset.get(cfg.key)
        if not group:
            continue
        lines.append(f"## {cfg.name}")
        lines.append("")
        lines.append(f"Data source: `{cfg.path}`")
        lines.append("")
        if cfg.documentation:
            lines.append("Supporting docs:")
            for doc in cfg.documentation:
                lines.append(f"- `{doc}`")
            lines.append("")
        for summary in group:
            lines.extend(_layer_section(summary))
        lines.append("")

    output_path.write_text("\n".join(lines), encoding="utf-8")


def _layer_section(summary: LayerSummary) -> List[str]:
    lines = [f"### Layer: {summary.display_name}", ""]
    lines.append(f"- Driver: `{summary.driver}`")
    if summary.geometry:
        lines.append(f"- Geometry type: `{summary.geometry}`")
    if summary.crs:
        lines.append(f"- CRS: `{summary.crs}`")
    if summary.feature_count is not None:
        lines.append(f"- Feature count: {summary.feature_count}")
    lines.append("")
    if summary.extra_lines:
        lines.append("Additional info:")
        for item in summary.extra_lines:
            lines.append(f"- {item}")
        lines.append("")
    if summary.fields is None:
        lines.append(
            "Shares schema with a previously listed layer; see the matching layer above for field definitions."
        )
        lines.append("")
        return lines
    lines.append("| Field | Type | Meaning |")
    lines.append("| --- | --- | --- |")
    for field in summary.fields:
        desc = field.description.replace("|", "/")
        lines.append(f"| `{field.name}` | {field.type_label} | {desc} |")
    lines.append("")
    return lines


# -- CLI ----------------------------------------------------------------------

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--dataset",
        dest="dataset_keys",
        action="append",
        help="Dataset key to include (default: all).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Markdown file to write (default: materials/dataset_field_inventory.md)",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    requested = set(args.dataset_keys or [])
    selected = [
        cfg for cfg in DATASETS if not requested or cfg.key in requested
    ]
    if not selected:
        print("No datasets selected.", file=sys.stderr)
        return 1
    summaries: List[LayerSummary] = []
    for cfg in selected:
        analyzer = DatasetAnalyzer(cfg)
        summaries.extend(analyzer.summarize())
    write_markdown_report(summaries, args.output)
    print(f"Report written to {args.output}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
